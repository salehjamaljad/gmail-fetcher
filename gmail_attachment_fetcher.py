import re
import base64
import zipfile
import io
from datetime import datetime, timedelta
from config import authenticate_gmail, upload_order_and_metadata
import openpyxl

def search_recent_emails(service):
    after_ts = int((datetime.utcnow() - timedelta(hours=2.1)).timestamp())
    query = (
        f'after:{after_ts} '
        'label:inbox has:attachment '
        '-from:me '
        '-from:osama@khodar.com '
        '(subject:"TMart Purchase Orders" OR subject:"Rabbit PO - Khodar trading and marketing" '
        'OR subject:"Khodar PO - Delivery Date" OR subject:"Khodar.com PO - Goodsmart" '
        'OR from:sherif.hossam@talabat.com OR from:rabbit.purchasing@rabbitmart.com '
        'OR from:abdelhamid.oraby@breadfast.com OR from:amir.maher@goodsmartegypt.com)'
        'OR from:Mohamed.OthmanAli@halan.com'
    )
    results = service.users().messages().list(userId='me', q=query).execute()
    return results.get('messages', [])

def extract_order_date_from_subject(subject):
    match = re.search(r"\[(\d{4}-\d{2}-\d{2})\]", subject)
    if match:
        return datetime.strptime(match.group(1), "%Y-%m-%d").date()
    return None

def get_next_delivery_date():
    tomorrow = datetime.today() + timedelta(days=1)
    if tomorrow.weekday() == 4:
        return (tomorrow + timedelta(days=1)).strftime("%Y-%m-%d")
    return tomorrow.strftime("%Y-%m-%d")

def determine_khateer_or_rabbit(xlsx_bytes):
    try:
        in_memory_file = io.BytesIO(xlsx_bytes)
        wb = openpyxl.load_workbook(in_memory_file, data_only=True)
        ws = wb.active
        value = str(ws["D10"].value).lower() if ws["D10"].value else ""
        return "Khateer" if "khateer" in value else "Rabbit"
    except Exception as e:
        print("Failed to inspect D10 for client check:", e)
        return "Rabbit"


def safe_zip_filename(filename: str) -> str:
    encoded = base64.urlsafe_b64encode(filename.encode()).decode()
    return f"{encoded}.zip"

def safe_xlsx_filename(filename: str) -> str:
    encoded = base64.urlsafe_b64encode(filename.encode()).decode()
    return f"{encoded}.xlsx"

def fetch_and_upload_orders():
    service = authenticate_gmail()
    messages = search_recent_emails(service)
    print(f"Found {len(messages)} matching emails")

    for idx, msg in enumerate(messages, 1):
        msg_data = service.users().messages().get(userId='me', id=msg['id'], format='full').execute()
        headers = msg_data.get("payload", {}).get("headers", [])
        subject = next((h["value"] for h in headers if h["name"] == "Subject"), "")
        sender = next((h["value"] for h in headers if h["name"] == "From"), "")
        snippet = msg_data.get("snippet", "")

        parts = msg_data["payload"].get("parts", [])
        zip_buffer = io.BytesIO()
        order_date = datetime.today().strftime("%Y-%m-%d")
        delivery_date = get_next_delivery_date()
        status = "Pending"
        client = "Unknown"
        city = None
        po_number = None

        # --- BreadFast ---
        if subject.lower().startswith("khodar po - delivery date") or "abdelhamid.oraby@breadfast.com" in sender.lower():
            client = "Breadfast"
            order_date = datetime.today().strftime("%Y-%m-%d")

            date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", subject)
            if date_match:
                delivery_date = datetime.strptime(date_match.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")
            else:
                delivery_date = get_next_delivery_date()

            subject_lower = subject.lower()
            if "alex" in subject_lower:
                city = "Alexandria"
            elif "mansoura" in subject_lower:
                city = "Mansoura"
            else:
                city_match = re.search(r"\((.*?)\)", subject)
                if city_match:
                    city = city_match.group(1).strip().capitalize()

            for part in parts:
                filename = part.get("filename")
                body = part.get("body", {})
                if filename and "attachmentId" in body and filename.lower().endswith(".pdf"):
                    att_id = body["attachmentId"]
                    attachment = service.users().messages().attachments().get(
                        userId='me', messageId=msg['id'], id=att_id).execute()
                    file_data = base64.urlsafe_b64decode(attachment['data'].encode("UTF-8"))

                    try:
                        upload_response = upload_order_and_metadata(
                            file_bytes=file_data,
                            filename=filename,
                            client=client,
                            order_type="Purchase Order",
                            order_date=order_date,
                            delivery_date=delivery_date,
                            status=status,
                            city=city,
                            po_number=None
                        )
                        print(f"  Uploaded BreadFast PDF. Supabase ID: {upload_response[0].get('id')}")
                    except Exception as e:
                        print(f"  BreadFast Upload failed: {e}")
            continue

        # --- GoodsMart ---
        if "amir.maher@goodsmartegypt.com" in sender.lower() or "khodar.com po - goodsmart" in subject.lower():
            client = "GoodsMart"
            order_date = datetime.today().strftime("%Y-%m-%d")
            status = "Pending"

            # Extract delivery date
            delivery_match = re.search(r"Expected Delivery Date:\s*(\d{1,2}/\d{1,2}/\d{4})", snippet)
            if delivery_match:
                delivery_date = datetime.strptime(delivery_match.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")

            # Extract PO Number
            po_match = re.search(r"PO No\s*(\d+)", snippet)
            if po_match:
                po_number = po_match.group(1)

            for part in parts:
                filename = part.get("filename")
                body = part.get("body", {})
                if filename and "attachmentId" in body and filename.lower().endswith(".xlsx"):
                    att_id = body["attachmentId"]
                    attachment = service.users().messages().attachments().get(
                        userId='me', messageId=msg['id'], id=att_id).execute()
                    file_data = base64.urlsafe_b64decode(attachment['data'].encode("UTF-8"))

                    try:
                        upload_response = upload_order_and_metadata(
                            file_bytes=file_data,
                            filename=filename,
                            client=client,
                            order_type="Purchase Order",
                            order_date=order_date,
                            delivery_date=delivery_date,
                            status=status,
                            city=None,
                            po_number=po_number
                        )
                        print(f"  Uploaded GoodsMart file. Supabase ID: {upload_response[0].get('id')}")
                    except Exception as e:
                        print(f"  GoodsMart upload failed: {e}")
            continue
        
                # --- Halan ---
        if "طلبيه الخضار شركة خضار دوت كوم -حالا" in subject or "Mohamed.OthmanAli@halan.com" in sender:
            client = "Halan"
            order_date = datetime.today().strftime("%Y-%m-%d")
            status = "Pending"

            # Compute next delivery date (nearest Saturday or Wednesday)
            today = datetime.today()
            weekday = today.weekday()  # Monday = 0, Sunday = 6

            if weekday in [5, 6, 0, 1]:  # Saturday, Sunday, Monday, Tuesday
                days_until_wed = (2 - weekday) % 7 or 7
                delivery_date = (today + timedelta(days=days_until_wed)).strftime("%Y-%m-%d")
            else:  # Wednesday, Thursday, Friday
                days_until_sat = (5 - weekday) % 7 or 7
                delivery_date = (today + timedelta(days=days_until_sat)).strftime("%Y-%m-%d")


            # Extract PO Number from body/snippet between 'مدينه نصر' and 'حدايق الاهرام'
            po_number = None
            body_text = snippet.replace("\n", " ")
            match = re.search(r"مدينه نصر(.*?)حدايق الاهرام", body_text)
            if match:
                po_number = match.group(1).strip()

            for part in parts:
                filename = part.get("filename")
                body = part.get("body", {})
                if filename and "attachmentId" in body and filename.lower().endswith(".xlsx"):
                    att_id = body["attachmentId"]
                    attachment = service.users().messages().attachments().get(
                        userId='me', messageId=msg['id'], id=att_id).execute()
                    file_data = base64.urlsafe_b64decode(attachment['data'].encode("UTF-8"))
                    filename = safe_xlsx_filename(filename)
                    try:
                        upload_response = upload_order_and_metadata(
                            file_bytes=file_data,
                            filename=filename,
                            client=client,
                            order_type="Purchase Order",
                            order_date=order_date,
                            delivery_date=delivery_date,
                            status=status,
                            city=None,
                            po_number=po_number
                        )
                        print(f"  Uploaded Halan file. Supabase ID: {upload_response[0].get('id')}")
                    except Exception as e:
                        print(f"  Halan upload failed: {e}")
            continue
        

        # --- General (Rabbit / Khateer / Talabat) ---
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for part in parts:
                filename = part.get("filename")
                body = part.get("body", {})
                if filename and "attachmentId" in body:
                    if not any(filename.lower().endswith(ext) for ext in ['.pdf', '.xls', '.xlsx', '.csv', '.zip']):
                        continue
                    att_id = body["attachmentId"]
                    attachment = service.users().messages().attachments().get(
                        userId='me', messageId=msg['id'], id=att_id).execute()
                    file_data = base64.urlsafe_b64decode(attachment['data'].encode("UTF-8"))
                    zipf.writestr(filename, file_data)

                    if filename.lower().endswith(".xlsx") and client == "Unknown":
                        client = determine_khateer_or_rabbit(file_data)

        zip_buffer.seek(0)
        zip_filename = f"{filename}.zip"
        zip_filename = safe_zip_filename(filename)
        zip_filename = client + zip_filename

        if subject.lower().startswith("tmart purchase orders") or "sherif.hossam@talabat.com" in sender.lower():
            match = extract_order_date_from_subject(subject)
            if not match:
                print(f"No valid date in subject for email {idx}, using today's date instead.")
                match = datetime.today()
            order_date = match.strftime("%Y-%m-%d")
            delivery_date = (match + timedelta(days=2)).strftime("%Y-%m-%d")
            client = "Talabat"

        print(f"\nProcessing email {idx}")
        print(f"  Subject: {subject}")
        print(f"  From: {sender}")
        print(f"  Order Date: {order_date}")
        print(f"  Delivery Date: {delivery_date}")
        print(f"  Client: {client}")

        try:
            upload_response = upload_order_and_metadata(
                file_bytes=zip_buffer.read(),
                filename=zip_filename,
                client=client,
                order_type="Purchase Order",
                order_date=order_date,
                delivery_date=delivery_date,
                status=status,
                city=city,
                po_number=po_number
            )
            print(f"  Uploaded successfully. Supabase ID: {upload_response[0].get('id')}")
        except Exception as e:
            print(f"  Upload failed: {e}")

if __name__ == '__main__':
    fetch_and_upload_orders()
